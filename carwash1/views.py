import imp
from operator import imod
import re
from django.shortcuts import render,redirect,get_object_or_404
from django.core.paginator import Paginator
from django.http import Http404
from .models import Ingreso
from django.contrib.auth.decorators import login_required,permission_required

#importar form

from .forms import CustomUserCreationForm, IngresoForm,UserCreationForm,ReservaForm,IngresoUpdateForm

from django.contrib.auth import authenticate,login
from .filters import FilterIngreso

from django.contrib import messages


def inicio(request):
    return render(request,'home.html')

from django.contrib import messages
@login_required
def ingreso(request):
    if request.method == 'POST':
        formulario = IngresoForm(data=request.POST, files=request.FILES)
        if formulario.is_valid():
            try:
                ingreso = formulario.save()
                messages.success(request, "Vehículo ingresado correctamente.")
                return redirect('tabla_ingresos')
            except Exception as e:
                messages.error(request, f'Error al guardar el ingreso: {str(e)}')
        else:
            # Mostrar errores del formulario sin el "Excelente"
            messages.warning(request, 'Formulario no válido. Corrija los errores.')
            return render(request, 'ingreso.html', {'form': formulario})
    else:
        formulario = IngresoForm()
    
    return render(request, 'ingreso.html', {'form': formulario})
def reserva(request):
    
    data = {
        'form':ReservaForm()
    }

    if request.method == 'POST':
        formulario = ReservaForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request,"Vehículo Reservado Correctamente")
        else:
            data["form"]=formulario
    return render(request,'reserva.html',data)


def contacto(request):
    return render(request,'contacto.html')

@login_required
def tabla_ingresos(request):
    data = Ingreso.objects.all()
    filters = FilterIngreso(request.GET, queryset=data)
    filtered_data = filters.qs
    
    paginator = Paginator(filtered_data, 10)  # 10 ingresos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        "entity": page_obj,
        "filters": filters,
        "paginator": paginator,
    }

    return render(request, 'tablaingresos.html', context)

@permission_required('carwash1.change_ingreso')
def modificar_ingreso(request, id):
    ingreso = get_object_or_404(Ingreso, id=id)
    
    if request.method == 'POST':
        formulario = IngresoUpdateForm(data=request.POST, instance=ingreso)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado Correctamente")
            return redirect('tabla_ingresos')  # Asegúrate de que 'tabla_ingresos' sea el nombre correcto de la URL
    else:
        formulario = IngresoUpdateForm(instance=ingreso)

    data = {
        'form': formulario
    }

    return render(request, 'modificaringreso.html', data)

@permission_required('carwash1.delete_ingreso')
def eliminar_ingreso(request,id):
    ingreso = get_object_or_404(Ingreso, id=id)
    ingreso.delete()
    messages.success(request,"Eliminado Correctamente")
    return redirect(to=tabla_ingresos)

from django.contrib.auth.models import Permission
@login_required
def registro(request):
    data = {
        'form': CustomUserCreationForm()
    }

    if request.method == 'POST':
        formulario = CustomUserCreationForm(data=request.POST)
        if formulario.is_valid():
            user = formulario.save()

            # Asignar permisos específicos
            permiso_ingreso = Permission.objects.get(codename='view_ingreso')
            permiso_tabla_ingresos = Permission.objects.get(codename='add_ingreso')
            user.user_permissions.add(permiso_ingreso, permiso_tabla_ingresos)
            
            # Autenticar y loguear al usuario
            login(request, user)
            messages.success(request, "Te has registrado correctamente")
            return redirect(to=inicio)
        data["form"] = formulario
        
    return render(request, 'registration/registro.html', data)


    
import openpyxl
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from datetime import datetime
import base64
from io import BytesIO
from PIL import Image as PILImage
import tempfile

def export_ingresos_to_excel(request):
    data = Ingreso.objects.all()
    filters = FilterIngreso(request.GET, queryset=data)
    filtered_data = filters.qs

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=ingresos.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Ingresos'

    columns = [
        'FECHA INGRESO', 'PATENTE', 'VEHÍCULO', 'LAVADO', 'TIPO DE PAGO', 'VALOR', 'PROPINA',
        'CONTACTO', 'ESTADO VEHÍCULO', 'TIPO DOCUMENTO', 'ESTADO FACTURA', 'RUT', 'CLIENTE ACEPTA',
        'FIRMA ELECTRÓNICA'
    ]

    # Write header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    # Write data
    for row_num, ing in enumerate(filtered_data, 2):
        row = [
            ing.fecha_ingreso.replace(tzinfo=None) if isinstance(ing.fecha_ingreso, datetime) else ing.fecha_ingreso,
            ing.patente,
            ing.vehiculo,
            ing.lavado,
            ing.tipo_de_pago,
            ing.valor,
            ing.propina,
            ing.contacto,
            ing.estado_vehiculo,
            ing.tipo_doc,
            ing.estado_factura,
            ing.Rut,
            'Sí' if ing.cliente_acepta else 'No',
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            if isinstance(cell_value, datetime):
                cell.value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
            else:
                cell.value = cell_value

        # Add signature image if exists
        if ing.firma_electronica:
            try:
                # Decode base64 string
                signature_data = base64.b64decode(ing.firma_electronica.split(',')[1])
                img = PILImage.open(BytesIO(signature_data))

                # Resize image
                img.thumbnail((60, 50))  # Adjust the size as necessary

                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
                    img.save(tmp_file.name)
                    excel_img = Image(tmp_file.name)
                    cell = worksheet.cell(row=row_num, column=len(row) + 1)
                    worksheet.add_image(excel_img, cell.coordinate)
            except Exception as e:
                cell = worksheet.cell(row=row_num, column=len(row) + 1)
                cell.value = f"Error: {str(e)}"
        else:
            cell = worksheet.cell(row=row_num, column=len(row) + 1)
            cell.value = "No disponible"

    workbook.save(response)
    return response